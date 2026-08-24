"""E7A: add the model's own column next to Sandeep's labels for manual review.

Writes results/e7a/e7a_labeling_review.xlsx containing:
  Round1 / Round2 sheets: original columns + human_label, model_pred,
          model_conf, match (yellow = disagree)
  Disagreements sheet:    only mismatching reviews, sorted by confidence

Run: venv\\Scripts\\python.exe -u src\\experiments\\review_e7a_excel.py
"""

import os
import sys

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
E7A_DIR = os.path.join(REPO_ROOT, "results", "e7a")
OUT_XLSX = os.path.join(E7A_DIR, "e7a_labeling_review.xlsx")
LABELS = ["negative", "neutral", "positive"]

YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="D9EAD3")
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)


def to_label(score: int) -> str:
    return "negative" if score <= 3 else ("positive" if score >= 7 else "neutral")


def load(sheet_csv):
    df = pd.read_csv(os.path.join(E7A_DIR, sheet_csv), encoding="utf-8-sig",
                     keep_default_na=False)
    key = pd.read_csv(os.path.join(E7A_DIR, "e7a_key.csv"))
    df["review_id"] = df["review_id"].astype(int)
    df["score"] = df["score"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    scored = df[df["score"] != ""].copy()
    scored["human_score"] = scored["score"].astype(int)
    scored["human_label"] = scored["human_score"].apply(to_label)
    scored = scored.merge(key[["review_id", "pred_label", "pred_conf"]],
                          on="review_id", how="left")
    scored["match"] = scored["human_label"] == scored["pred_label"]
    return df, scored


def style_sheet(ws, n_cols, wrap_col=3):
    ws.freeze_panes = "A2"
    widths = [10, 100, 8, 14, 12, 12, 11, 9]
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i <= len(widths) else 14
        c = ws.cell(row=1, column=i)
        c.fill = HDR_FILL
        c.font = HDR_FONT
    for row in ws.iter_rows(min_row=2, min_col=wrap_col, max_col=wrap_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_sheet(wb, name, scored, fill):
    ws = wb.create_sheet(name)
    cols = ["review_id", "Review Text", "score", "comment",
            "human_label", "model_pred", "model_conf", "match"]
    ws.append(cols)
    for _, r in scored.iterrows():
        ws.append([int(r["review_id"]), r["Review Text"], r["human_score"],
                   r.get("comment", ""), r["human_label"], r["pred_label"],
                   round(float(r["pred_conf"]), 3),
                   "OK" if r["match"] else "DIFF"])
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(cols)):
        if row[7].value == "DIFF":
            for c in row:
                c.fill = fill
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=(c.column == 2))
    style_sheet(ws, len(cols), wrap_col=2)
    return ws


def main():
    _, s1 = load("e7a_labeling_round1.csv")
    _, s2 = load("e7a_labeling_round2.csv")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xl:
        wb = xl.book
        write_sheet(wb, "Round1", s1.sort_values("match"), YELLOW)
        write_sheet(wb, "Round2", s2.sort_values("match"), YELLOW)

        dis = pd.concat([s1, s2]).loc[lambda d: ~d["match"]].copy()
        dis = dis.sort_values("pred_conf", ascending=False)
        ws = wb.create_sheet("Disagreements")
        ws.append(["review_id", "Review Text", "your_score", "your_label",
                   "model_pred", "model_conf", "round"])
        rounds = pd.concat([
            s1.assign(round="R1"), s2.assign(round="R2")])
        dis = rounds.loc[~rounds["match"]].sort_values("pred_conf", ascending=False)
        for _, r in dis.iterrows():
            ws.append([int(r["review_id"]), r["Review Text"], r["human_score"],
                       r["human_label"], r["pred_label"],
                       round(float(r["pred_conf"]), 3), r["round"]])
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=7):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=(c.column == 2))
                c.fill = GREEN
        style_sheet(ws, 7, wrap_col=2)

    n1, ok1 = len(s1), int(s1["match"].sum())
    n2, ok2 = len(s2), int(s2["match"].sum())
    print(f"Round1: {n1} labeled | you-vs-model agree {ok1} ({ok1/n1:.1%}) | diff {n1-ok1}")
    print(f"Round2: {n2} labeled | agree {ok2} ({ok2/n2:.1%}) | diff {n2-ok2}")
    print(f"disagreements to review: {len(dis)}")
    print(f"-> {os.path.relpath(OUT_XLSX, REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
